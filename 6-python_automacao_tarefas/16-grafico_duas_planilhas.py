from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference, Series

dict_ano = {}

# 1 - Importando despesas
arquivo1 = load_workbook(filename='files/despesas.xlsx')
ws1 = arquivo1['despesas']
max_linhas = ws1.max_row

for i in range(2, max_linhas + 1):
    # print(ws1['A%s' %i].value)
    # print(ws1['B%s' %i].value)
    dict_ano[ws1['A%s' %i].value] = {'despesa':ws1['B%s' %i].value, 'receita':0}
    
# print(dict_ano)

# 2 - Importando receitas
arquivo2 = load_workbook(filename='files/receitas.xlsx')
ws2 = arquivo2['receitas']
max_linhas = ws2.max_row

for i in range(2, max_linhas + 1):
    dict_ano[ws2['A%s' %i].value]['receita'] = ws2['B%s' %i].value
    
# print(dict_ano)

# 3 - Criando planilha
wb = Workbook()
ws = wb.active

ws['A1'] = 'Ano'
ws['B1'] = 'Despesa'
ws['C1'] = 'Receita'

i = 2
for key, value in dict_ano.items():
    ws['A%s' %i] = key
    ws['B%s' %i] = value['despesa']
    ws['C%s' %i] = value['receita']
    i += 1
    
# Criação do grafico
chart1 = BarChart()
chart1.type = 'col'
chart1.style = 12
chart1.title = 'Receita x Despesa por ano'
chart1.y_axis.title = 'R$'
chart1.x_axis.title = 'Ano'
    
# Criação das referências
data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=i)
anos = Reference(ws, min_col=1, min_row=2, max_row=i)

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(anos)
chart1.shape = 4
ws.add_chart(chart1, 'A%s' %(i+2))
    
wb.save(filename='files/demonstrativo.xlsx')