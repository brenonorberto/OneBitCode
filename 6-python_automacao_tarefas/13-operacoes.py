from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(filename='files/gastos.xlsx')
planilha = wb['Sheet']

# print(planilha)

valor_total = 0
# 1 - Somando valores da planilha
for i in range(2, planilha.max_row + 1):
    valor = int(planilha['B%s' %i].value)
    valor_total += valor # valor_total = valor_total + valor

# print('Total de gastos: R${0}'.format(valor_total))
planilha['A10'] = 'Total de gastos'
planilha['B10'] = valor_total
# wb.save(filename='files/gastos.xlsx')

# 2 - Mesclar celulas
planilha['A11'] = 'Teste'
planilha.merge_cells('A11:B11')
planilha.unmerge_cells('A11:B11')
# wb.save(filename='files/gastos.xlsx')

# 3 - Inserindo imagem
img = Image('BBAS3.png')
planilha.add_image(img, 'A12')

wb.save(filename='files/gastos.xlsx')

# 4 - Deletando celulas
planilha.delete_rows(1)
planilha.delete_cols(3)
# wb.save(filename='files/gastos2.xlsx')


